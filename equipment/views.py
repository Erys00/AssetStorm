from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from django.http import HttpResponse, Http404
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from .models import Equipment, EquipmentTransfer
from .forms import EquipmentForm, EquipmentTransferForm, CustomLoginForm, EquipmentTransferApprovalForm
from .decorators import it_staff_required, can_access_equipment, get_user_equipment_queryset, can_transfer_equipment


def custom_login(request):
    """Niestandardowy widok logowania"""
    if request.user.is_authenticated:
        return redirect('equipment:equipment_list')
    
    if request.method == 'POST':
        form = CustomLoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Witaj, {user.get_full_name() or user.username}!')
                return redirect('equipment:equipment_list')
            else:
                messages.error(request, 'Nieprawidłowa nazwa użytkownika lub hasło.')
        else:
            messages.error(request, 'Nieprawidłowa nazwa użytkownika lub hasło.')
    else:
        form = CustomLoginForm()
    
    return render(request, 'equipment/login.html', {'form': form})


def custom_logout(request):
    """Niestandardowy widok wylogowania"""
    if request.user.is_authenticated:
        username = request.user.username
        logout(request)
        messages.success(request, f'Użytkownik {username} został wylogowany.')
    return redirect('equipment:login')


@login_required
@can_access_equipment
def dashboard(request):
    """Dashboard z analityką sprzętu"""
    # Pobierz sprzęt dostępny dla użytkownika
    equipment_list = get_user_equipment_queryset(request.user)
    
    # Statystyki podstawowe
    total_equipment = equipment_list.count()
    available_count = equipment_list.filter(status='available').count()
    in_use_count = equipment_list.filter(status='in_use').count()
    service_count = equipment_list.filter(status='service').count()
    retired_count = equipment_list.filter(status='retired').count()
    
    # Statystyki finansowe (tylko dla IT/admin)
    total_value = 0
    avg_value = 0
    if request.user.is_superuser or request.user.groups.filter(name='IT').exists():
        total_value = equipment_list.aggregate(
            total=Sum('purchase_price')
        )['total'] or 0
        avg_value = equipment_list.aggregate(
            avg=Avg('purchase_price')
        )['avg'] or 0
    
    # Statystyki po typach sprzętu
    equipment_by_type = equipment_list.values('type').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Statystyki po dostawcach
    equipment_by_supplier = equipment_list.values('supplier').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Alerty gwarancyjne (tylko dla IT/admin)
    warranty_alerts = []
    if request.user.is_superuser or request.user.groups.filter(name='IT').exists():
        # Sprzęt z gwarancją kończącą się w ciągu 30 dni
        thirty_days_from_now = timezone.now().date() + timedelta(days=30)
        warranty_alerts = equipment_list.filter(
            warranty_end_date__lte=thirty_days_from_now,
            warranty_end_date__gte=timezone.now().date()
        ).order_by('warranty_end_date')[:10]
    
    # Ostatnie transfery (tylko dla IT/admin)
    recent_transfers = []
    if request.user.is_superuser or request.user.groups.filter(name='IT').exists():
        recent_transfers = EquipmentTransfer.objects.select_related(
            'equipment', 'from_user', 'to_user', 'transferred_by'
        ).order_by('-transfer_date')[:5]
    
    # Sprzęt w serwisie (długo)
    long_service = equipment_list.filter(status='service').order_by('updated_at')[:5]
    
    context = {
        'total_equipment': total_equipment,
        'available_count': available_count,
        'in_use_count': in_use_count,
        'service_count': service_count,
        'retired_count': retired_count,
        'total_value': total_value,
        'avg_value': avg_value,
        'equipment_by_type': equipment_by_type,
        'equipment_by_supplier': equipment_by_supplier,
        'warranty_alerts': warranty_alerts,
        'recent_transfers': recent_transfers,
        'long_service': long_service,
        'is_it_staff': request.user.is_superuser or request.user.groups.filter(name='IT').exists(),
    }
    
    return render(request, 'equipment/dashboard.html', context)


@login_required
def equipment_list(request):
    """Lista wszystkich sprzętów z filtrowaniem i wyszukiwaniem"""
    # Użyj funkcji pomocniczej do filtrowania sprzętu
    equipment_list = get_user_equipment_queryset(request.user)
    
    # Wyszukiwanie
    search_query = request.GET.get('search', '')
    if search_query:
        # Sprawdź czy wyszukiwany tekst to liczba (ID)
        search_filters = Q(name__icontains=search_query) | \
                        Q(serial_number__icontains=search_query) | \
                        Q(type__icontains=search_query) | \
                        Q(invoice_number__icontains=search_query) | \
                        Q(location__icontains=search_query) | \
                        Q(supplier__icontains=search_query) | \
                        Q(notes__icontains=search_query)
        
        # Dodaj wyszukiwanie po ID jeśli tekst to liczba
        try:
            search_id = int(search_query)
            search_filters |= Q(id=search_id)
        except ValueError:
            # Jeśli nie można przekonwertować na int, pomiń wyszukiwanie po ID
            pass
        
        equipment_list = equipment_list.filter(search_filters)
    
    # Filtrowanie po statusie
    status_filter = request.GET.get('status', '')
    if status_filter:
        equipment_list = equipment_list.filter(status=status_filter)
    
    # Filtrowanie po użytkowniku (tylko dla IT)
    user_filter = request.GET.get('user', '')
    if user_filter and (request.user.is_superuser or request.user.groups.filter(name='IT').exists()):
        equipment_list = equipment_list.filter(
            Q(assigned_to__username__icontains=user_filter) |
            Q(assigned_to__first_name__icontains=user_filter) |
            Q(assigned_to__last_name__icontains=user_filter)
        )
    
    # Filtrowanie po lokalizacji
    location_filter = request.GET.get('location', '')
    if location_filter:
        equipment_list = equipment_list.filter(location__icontains=location_filter)
    
    # Filtrowanie po dostawcy
    supplier_filter = request.GET.get('supplier', '')
    if supplier_filter:
        equipment_list = equipment_list.filter(supplier__icontains=supplier_filter)
    
    # Paginacja
    paginator = Paginator(equipment_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'user_filter': user_filter,
        'location_filter': location_filter,
        'supplier_filter': supplier_filter,
        'status_choices': Equipment.STATUS_CHOICES,
        'is_it_staff': request.user.is_superuser or request.user.groups.filter(name='IT').exists(),
    }
    return render(request, 'equipment/equipment_list.html', context)


@login_required
@can_access_equipment
def equipment_detail(request, pk):
    """Szczegóły sprzętu"""
    equipment = get_object_or_404(Equipment, pk=pk)
    
    # Obsługa generowania QR kodu
    if request.method == 'POST' and request.headers.get('Content-Type') == 'application/json':
        import json
        data = json.loads(request.body)
        if data.get('action') == 'generate_qr':
            if request.user.is_superuser or request.user.groups.filter(name='IT').exists():
                equipment.generate_qr_code()
                return HttpResponse(json.dumps({'success': True}), content_type='application/json')
            else:
                return HttpResponse(json.dumps({'success': False, 'error': 'Brak uprawnień'}), content_type='application/json')
    
    transfer_history = equipment.get_transfer_history()
    
    context = {
        'equipment': equipment,
        'transfer_history': transfer_history,
        'can_transfer': can_transfer_equipment(request.user, equipment),
        'is_it_staff': request.user.is_superuser or request.user.groups.filter(name='IT').exists(),
    }
    return render(request, 'equipment/equipment_detail.html', context)


@login_required
@it_staff_required
def equipment_create(request):
    """Dodawanie nowego sprzętu"""
    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            equipment = form.save()
            messages.success(request, f'Sprzęt "{equipment.name}" został dodany.')
            return redirect('equipment:equipment_detail', pk=equipment.pk)
    else:
        form = EquipmentForm()
    
    return render(request, 'equipment/equipment_form.html', {'form': form, 'title': 'Dodaj sprzęt'})


@login_required
@can_access_equipment
def equipment_update(request, pk):
    """Edycja sprzętu"""
    equipment = get_object_or_404(Equipment, pk=pk)
    
    # Sprawdź czy użytkownik może edytować sprzęt
    if not (request.user.is_superuser or request.user.groups.filter(name='IT').exists()):
        if equipment.assigned_to != request.user:
            messages.error(request, 'Nie możesz edytować tego sprzętu.')
            return redirect('equipment:my_equipment')
    
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            equipment = form.save()
            messages.success(request, f'Sprzęt "{equipment.name}" został zaktualizowany.')
            return redirect('equipment:equipment_detail', pk=equipment.pk)
    else:
        form = EquipmentForm(instance=equipment)
    
    return render(request, 'equipment/equipment_form.html', {
        'form': form, 
        'equipment': equipment,
        'title': 'Edytuj sprzęt'
    })


@login_required
@it_staff_required
def equipment_delete(request, pk):
    """Usuwanie sprzętu"""
    equipment = get_object_or_404(Equipment, pk=pk)
    if request.method == 'POST':
        equipment_name = equipment.name
        equipment.delete()
        messages.success(request, f'Sprzęt "{equipment_name}" został usunięty.')
        return redirect('equipment:equipment_list')
    
    return render(request, 'equipment/equipment_confirm_delete.html', {'equipment': equipment})


@login_required
@it_staff_required
def equipment_transfer(request, pk):
    """Przekazywanie sprzętu innemu użytkownikowi - tylko dla IT"""
    equipment = get_object_or_404(Equipment, pk=pk)
    
    # Sprawdź czy użytkownik może przekazać sprzęt
    if not can_transfer_equipment(request.user, equipment):
        messages.error(request, 'Nie możesz przekazać tego sprzętu.')
        return redirect('equipment:equipment_detail', pk=equipment.pk)
    
    if request.method == 'POST':
        form = EquipmentTransferForm(request.POST)
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.equipment = equipment
            transfer.from_user = equipment.assigned_to
            transfer.transferred_by = request.user
            transfer.approval_status = 'pending'  # Nowy transfer czeka na akceptację
            
            # Sprawdź czy to_user jest poprawnie ustawiony
            if not transfer.to_user:
                messages.error(request, 'Musisz wybrać użytkownika, do którego chcesz przekazać sprzęt.')
                return render(request, 'equipment/equipment_transfer.html', {
                    'form': form,
                    'equipment': equipment,
                    'title': f'Przekaż sprzęt: {equipment.name}'
                })
            
            # Zapisz transfer (bez aktualizacji sprzętu - to się stanie po akceptacji)
            transfer.save()
            
            to_user_name = transfer.to_user.get_full_name() or transfer.to_user.username
            
            messages.success(
                request, 
                f'Wniosek o przekazanie sprzętu "{equipment.name}" został wysłany do użytkownika {to_user_name}. '
                f'Przekazanie zostanie zrealizowane po akceptacji przez odbiorcę.'
            )
            return redirect('equipment:equipment_detail', pk=equipment.pk)
        else:
            messages.error(request, 'Formularz zawiera błędy. Sprawdź wprowadzone dane.')
    else:
        form = EquipmentTransferForm()
    
    context = {
        'form': form,
        'equipment': equipment,
        'title': f'Przekaż sprzęt: {equipment.name}'
    }
    return render(request, 'equipment/equipment_transfer.html', context)


@login_required
def my_equipment(request):
    """Lista sprzętu przypisanego do zalogowanego użytkownika"""
    equipment_list = get_user_equipment_queryset(request.user)
    
    # Wyszukiwanie
    search_query = request.GET.get('search', '')
    if search_query:
        # Sprawdź czy wyszukiwany tekst to liczba (ID)
        search_filters = Q(name__icontains=search_query) | \
                        Q(serial_number__icontains=search_query) | \
                        Q(type__icontains=search_query) | \
                        Q(invoice_number__icontains=search_query) | \
                        Q(location__icontains=search_query) | \
                        Q(supplier__icontains=search_query) | \
                        Q(notes__icontains=search_query)
        
        # Dodaj wyszukiwanie po ID jeśli tekst to liczba
        try:
            search_id = int(search_query)
            search_filters |= Q(id=search_id)
        except ValueError:
            # Jeśli nie można przekonwertować na int, pomiń wyszukiwanie po ID
            pass
        
        equipment_list = equipment_list.filter(search_filters)
    
    # Filtrowanie po statusie
    status_filter = request.GET.get('status', '')
    if status_filter:
        equipment_list = equipment_list.filter(status=status_filter)
    
    # Paginacja
    paginator = Paginator(equipment_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': Equipment.STATUS_CHOICES,
        'is_it_staff': request.user.is_superuser or request.user.groups.filter(name='IT').exists(),
    }
    return render(request, 'equipment/my_equipment.html', context)


@login_required
@it_staff_required
def export_equipment_excel(request):
    """Eksport sprzętu do pliku Excel"""
    # Pobierz sprzęt
    equipment_list = Equipment.objects.all().order_by('name')
    
    # Utwórz nowy skoroszyt
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sprzęt IT"
    
    # Nagłówki
    headers = [
        'Nazwa', 'Typ', 'Numer seryjny', 'Numer faktury', 'Data zakupu',
        'Cena zakupu', 'Lokalizacja', 'Dostawca', 'Koniec gwarancji',
        'Status', 'Przypisany do', 'Uwagi', 'Data utworzenia'
    ]
    
    # Styl nagłówków
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Dodaj nagłówki
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dodaj dane
    for row, equipment in enumerate(equipment_list, 2):
        ws.cell(row=row, column=1, value=equipment.name)
        ws.cell(row=row, column=2, value=equipment.type)
        ws.cell(row=row, column=3, value=equipment.serial_number)
        ws.cell(row=row, column=4, value=equipment.invoice_number or '')
        ws.cell(row=row, column=5, value=equipment.purchase_date.strftime('%d.%m.%Y') if equipment.purchase_date else '')
        ws.cell(row=row, column=6, value=float(equipment.purchase_price) if equipment.purchase_price else '')
        ws.cell(row=row, column=7, value=equipment.location)
        ws.cell(row=row, column=8, value=equipment.supplier or '')
        ws.cell(row=row, column=9, value=equipment.warranty_end_date.strftime('%d.%m.%Y') if equipment.warranty_end_date else '')
        ws.cell(row=row, column=10, value=equipment.get_status_display())
        ws.cell(row=row, column=11, value=equipment.assigned_to.get_full_name() if equipment.assigned_to else '')
        ws.cell(row=row, column=12, value=equipment.notes or '')
        ws.cell(row=row, column=13, value=equipment.created_at.strftime('%d.%m.%Y %H:%M'))
    
    # Dostosuj szerokość kolumn
    column_widths = [25, 20, 20, 15, 12, 12, 15, 15, 12, 12, 20, 30, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Utwórz odpowiedź HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="sprzet_it_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    # Zapisz skoroszyt
    wb.save(response)
    return response


@login_required
@it_staff_required
def export_transfers_excel(request):
    """Eksport transferów do pliku Excel"""
    # Pobierz transfery
    transfers = EquipmentTransfer.objects.select_related(
        'equipment', 'from_user', 'to_user', 'transferred_by'
    ).order_by('-transfer_date')
    
    # Utwórz nowy skoroszyt
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transfery sprzętu"
    
    # Nagłówki
    headers = [
        'Sprzęt', 'Numer seryjny', 'Od użytkownika', 'Do użytkownika',
        'Data transferu', 'Przekazane przez', 'Powód'
    ]
    
    # Styl nagłówków
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Dodaj nagłówki
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dodaj dane
    for row, transfer in enumerate(transfers, 2):
        ws.cell(row=row, column=1, value=transfer.equipment.name)
        ws.cell(row=row, column=2, value=transfer.equipment.serial_number)
        ws.cell(row=row, column=3, value=transfer.from_user.get_full_name() if transfer.from_user else 'System')
        ws.cell(row=row, column=4, value=transfer.to_user.get_full_name() if transfer.to_user else 'System')
        ws.cell(row=row, column=5, value=transfer.transfer_date.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=6, value=transfer.transferred_by.get_full_name() if transfer.transferred_by else 'System')
        ws.cell(row=row, column=7, value=transfer.reason or '')
    
    # Dostosuj szerokość kolumn
    column_widths = [25, 20, 20, 20, 18, 20, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # Utwórz odpowiedź HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="transfery_sprzetu_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    # Zapisz skoroszyt
    wb.save(response)
    return response


@login_required
def pending_transfers(request):
    """Lista przekazań oczekujących na akceptację przez użytkownika"""
    pending_transfers = EquipmentTransfer.objects.filter(
        to_user=request.user,
        approval_status='pending'
    ).select_related('equipment', 'from_user', 'transferred_by').order_by('-transfer_date')
    
    context = {
        'pending_transfers': pending_transfers,
    }
    return render(request, 'equipment/pending_transfers.html', context)


@login_required
def approve_transfer(request, transfer_id):
    """Akceptacja lub odrzucenie przekazania sprzętu"""
    transfer = get_object_or_404(EquipmentTransfer, pk=transfer_id)
    
    # Sprawdź czy użytkownik może zaakceptować to przekazanie
    if not transfer.can_be_approved_by(request.user):
        raise Http404("Nie możesz zaakceptować tego przekazania.")
    
    if request.method == 'POST':
        form = EquipmentTransferApprovalForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            reason = form.cleaned_data['reason']
            
            if action == 'approve':
                transfer.approve(request.user)
                messages.success(
                    request,
                    f'Przekazanie sprzętu "{transfer.equipment.name}" zostało zaakceptowane. '
                    f'Protokół przekazania został wygenerowany.'
                )
                return redirect('equipment:transfer_protocol', transfer_id=transfer.id)
            else:  # reject
                transfer.reject(request.user, reason)
                messages.success(
                    request,
                    f'Przekazanie sprzętu "{transfer.equipment.name}" zostało odrzucone.'
                )
                return redirect('equipment:pending_transfers')
    else:
        form = EquipmentTransferApprovalForm()
    
    context = {
        'transfer': transfer,
        'form': form,
    }
    return render(request, 'equipment/approve_transfer.html', context)


@login_required
def transfer_protocol(request, transfer_id):
    """Generowanie protokołu przekazania sprzętu w PDF"""
    transfer = get_object_or_404(EquipmentTransfer, pk=transfer_id)
    
    # Sprawdź czy użytkownik może zobaczyć protokół
    if not (request.user == transfer.to_user or 
            request.user.is_superuser or 
            request.user.groups.filter(name='IT').exists()):
        raise Http404("Nie masz uprawnień do wyświetlenia tego protokołu.")
    
    # Utwórz odpowiedź HTTP dla PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="protokol_przekazania_{transfer.equipment.serial_number}.pdf"'
    
    # Użyj czcionek z obsługą polskich znaków
    # ReportLab ma wbudowane czcionki z Unicode
    font_name = 'Helvetica'
    font_bold = 'Helvetica-Bold'
    
    # Alternatywnie można użyć Times-Roman lub Courier
    # font_name = 'Times-Roman'
    # font_bold = 'Times-Bold'
    
    # Utwórz dokument PDF z marginesami
    doc = SimpleDocTemplate(
        response, 
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    story = []
    
    # Style z polskimi znakami
    title_style = ParagraphStyle(
        'CustomTitle',
        fontName=font_bold,
        fontSize=14,
        spaceAfter=15,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        fontName=font_bold,
        fontSize=10,
        spaceAfter=8,
        spaceBefore=8,
        textColor=colors.darkblue
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        fontName=font_name,
        fontSize=9,
        spaceAfter=4,
        alignment=TA_LEFT
    )
    
    # Tytuł
    story.append(Paragraph("PROTOKÓŁ PRZEKAZANIA SPRZĘTU IT", title_style))
    story.append(Spacer(1, 10))
    
    # Dane przekazania w jednej tabeli
    basic_data = [
        ['Data przekazania:', transfer.transfer_date.strftime('%d.%m.%Y %H:%M')],
        ['Numer protokołu:', f'PP/{transfer.id:04d}/{transfer.transfer_date.year}'],
        ['Sprzęt:', transfer.equipment.name],
        ['Numer seryjny:', transfer.equipment.serial_number],
        ['Typ sprzętu:', transfer.equipment.type],
        ['Od użytkownika:', transfer.from_user.get_full_name() if transfer.from_user else 'System'],
        ['Do użytkownika:', transfer.to_user.get_full_name() if transfer.to_user else 'System'],
        ['Przekazane przez:', transfer.transferred_by.get_full_name() if transfer.transferred_by else 'System'],
        ['Lokalizacja:', transfer.equipment.location],
        ['Dostawca:', transfer.equipment.supplier or 'Nie podano'],
        ['Data zakupu:', transfer.equipment.purchase_date.strftime('%d.%m.%Y') if transfer.equipment.purchase_date else 'Nie podano'],
        ['Cena zakupu:', f'{transfer.equipment.purchase_price} PLN' if transfer.equipment.purchase_price else 'Nie podano'],
        ['Koniec gwarancji:', transfer.equipment.warranty_end_date.strftime('%d.%m.%Y') if transfer.equipment.warranty_end_date else 'Nie podano'],
        ['Status:', transfer.equipment.get_status_display()],
    ]
    
    basic_table = Table(basic_data, colWidths=[3.5*cm, 10*cm])
    basic_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('FONTNAME', (0, 0), (0, -1), font_bold),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    
    story.append(basic_table)
    story.append(Spacer(1, 10))
    
    # Powód przekazania (jeśli istnieje)
    if transfer.reason:
        story.append(Paragraph("Powód przekazania:", heading_style))
        story.append(Paragraph(transfer.reason, normal_style))
        story.append(Spacer(1, 8))
    
    # Status akceptacji
    status_text = f"Zaakceptowane przez: {transfer.approved_by.get_full_name() if transfer.approved_by else 'System'}<br/>"
    status_text += f"Data akceptacji: {transfer.approved_at.strftime('%d.%m.%Y %H:%M') if transfer.approved_at else 'Nie zaakceptowano'}"
    story.append(Paragraph("Status akceptacji:", heading_style))
    story.append(Paragraph(status_text, normal_style))
    story.append(Spacer(1, 15))
    
    # Podpisy w jednej linii
    signatures_data = [
        ['Odbiorca sprzętu:', '', 'Data:', ''],
        ['', '', '', ''],
        ['Podpis:', '', 'Podpis:', ''],
        ['', '', '', ''],
    ]
    
    signatures_table = Table(signatures_data, colWidths=[3*cm, 4*cm, 2*cm, 4*cm])
    signatures_table.setStyle(TableStyle([
        ('LINEBELOW', (1, 1), (1, 1), 1, colors.black),
        ('LINEBELOW', (1, 3), (1, 3), 1, colors.black),
        ('LINEBELOW', (3, 1), (3, 1), 1, colors.black),
        ('LINEBELOW', (3, 3), (3, 3), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(signatures_table)
    story.append(Spacer(1, 10))
    
    # Stopka
    story.append(Paragraph(
        f"Protokół wygenerowany automatycznie przez system AssetStorm • "
        f"Data wygenerowania: {timezone.now().strftime('%d.%m.%Y %H:%M')}",
        ParagraphStyle('Footer', fontName=font_name, fontSize=7, alignment=TA_CENTER)
    ))
    
    # Zbuduj PDF
    doc.build(story)
    return response
